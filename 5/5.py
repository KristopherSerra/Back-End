'''
5 - Un script Python que sea capaz de crear una planilla de cálculo a modo de reporte con
los datos de una campaña publicitaria finalizada dada. Este script, además debe enviar
por mail el reporte creado al cliente en cuestión.
'''

from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart3D,Reference, PieChart,ProjectedPieChart
 


informes = pd.read_excel("c:/Users/Kristopher/Desktop/De Todo/Universidad/LdL/BackEnd/5\BDE.xlsx")
tabla_pivote = informes.pivot_table(index="LOCALIDAD",columns='MODALIDAD', values='BLOQUE')
tabla_pivote.to_excel('c:/Users/Kristopher/Desktop/De Todo/Universidad/LdL/BackEnd/5\InformeAutomatico.xlsx', startrow=4,sheet_name='Octubre 2022')

lectura = load_workbook('c:/Users/Kristopher/Desktop/De Todo/Universidad/LdL/BackEnd/5\InformeAutomatico.xlsx')
pestaña = lectura['Octubre 2022']
min_columna= lectura.active.min_column
max_columna= lectura.active.max_column
min_fila= lectura.active.min_row
max_fila= lectura.active.max_row

graficoDeBarras = BarChart3D()
data = Reference(pestaña, min_col=min_columna,max_col=max_columna, min_row=min_fila,max_row=max_fila)
categorias = Reference(pestaña, min_col=min_columna,max_col=min_columna, min_row=min_fila,max_row=max_fila)

graficoDeBarras.add_data(data, titles_from_data=True)
graficoDeBarras.set_categories(categorias)
graficoDeBarras.title = "Resumen"
graficoDeBarras.y_axis.title = 'Modalidad'
graficoDeBarras.x_axis.title = 'Localidad'
pestaña.add_chart(graficoDeBarras)

'''
graficoDeTorta = PieChart()
data = Reference(pestaña, min_col=min_columna,max_col=max_columna, min_row=min_fila,max_row=max_fila)
categorias = Reference(pestaña, min_col=min_columna,max_col=min_columna, min_row=min_fila,max_row=max_fila)

graficoDeTorta.add_data(data, titles_from_data=True)
graficoDeTorta.set_categories(categorias)
graficoDeTorta.title = "Resumen"
pestaña.add_chart(graficoDeTorta)
'''

lectura.save('c:/Users/Kristopher/Desktop/De Todo/Universidad/LdL/BackEnd/5\InformeAutomatico.xlsx')